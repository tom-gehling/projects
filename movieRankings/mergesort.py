from openpyxl import load_workbook

def merge(left,right):
    result = []
    left_ptr = right_ptr = 0
    while left_ptr < len(left) and right_ptr < len(right):
        answer = input(f"'{left[left_ptr]}' or '{right[right_ptr]}' ?: ").capitalize()
        if answer == '1':
            result.append(left[left_ptr])
            left_ptr += 1
        elif answer == '2':
            result.append(right[right_ptr])
            right_ptr += 1
        else: 
            print('Error')
    result.extend(left[left_ptr:])
    result.extend(right[right_ptr:])

    return result

def merge_sort(arr):
    if len(arr) <= 1:
        return arr
    midpoint = int(len(arr)/2)
    left, right = merge_sort(arr[:midpoint]), merge_sort(arr[midpoint:])
    return merge(left, right)


def printRankings(arr):
    num = 1
    for i in arr:
        print(f'{num}. {i}')
        num+=1

def extract_movies(file):
    result = []
    mov = load_workbook(file)
    movies = mov.active
    length = movies.max_row
    for i in range(1,length):
        result.append(movies['A'+str(i)].value)
    return result


arr = extract_movies('movies.xlsx')
arr = merge_sort(arr)
printRankings(arr)